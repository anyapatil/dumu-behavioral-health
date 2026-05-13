import csv
import io
import logging
import os
import threading
from datetime import date, datetime, timedelta

from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    url_for,
)
from sqlalchemy import func

from models import (
    CHANNELS,
    COMPANY_TYPES,
    DELIVERY_SETTINGS,
    FUNDING_STATUSES,
    OUTCOMES,
    PIPELINE_STATUSES,
    PRIORITY_STATES,
    REVENUE_ESTIMATES,
    Company,
    Contact,
    OutreachLog,
    db,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ── App factory ──────────────────────────────────────────────────────────────

def create_app():
    app = Flask(__name__)

    default_db = "sqlite:////data/bh.db" if os.path.isdir("/data") else f"sqlite:///{os.path.join(os.getcwd(), 'bh.db')}"
    db_path = os.environ.get("DATABASE_URL", default_db)
    if db_path.startswith("postgres://"):
        db_path = db_path.replace("postgres://", "postgresql://", 1)

    app.config["SQLALCHEMY_DATABASE_URI"]        = db_path
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.secret_key = os.environ.get("SECRET_KEY", "dumu-bh-secret-2025")

    db.init_app(app)

    with app.app_context():
        db.create_all()
        _seed_from_csv()

    return app


def _seed_from_csv():
    if Company.query.count() > 0:
        return
    csv_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "outputs", "bh_companies.csv"
    )
    if not os.path.exists(csv_path):
        return

    TYPE_MAP = {t: t for t, _ in COMPANY_TYPES}
    count = 0
    with open(csv_path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            name = (row.get("company_name") or "").strip()
            if not name:
                continue
            db.session.add(Company(
                company_name            = name,
                website_url             = row.get("website_url") or None,
                phone                   = row.get("phone") or None,
                email                   = row.get("email") or None,
                address                 = row.get("address") or None,
                city                    = row.get("city") or None,
                state                   = row.get("state") or None,
                zip                     = row.get("zip") or None,
                description             = row.get("description") or None,
                company_type            = TYPE_MAP.get(row.get("company_type", ""), "behavioral_health"),
                delivery_setting        = row.get("delivery_setting") or "multi_setting",
                revenue_estimate        = row.get("revenue_estimate") or "unknown",
                funding_status          = row.get("funding_status") or "unknown",
                rank_score              = int(row.get("rank_score") or 50),
                status                  = row.get("status") or "uncontacted",
                verified                = (row.get("verified") or "").lower() == "yes",
                positive_signals        = row.get("positive_signals") or None,
                notes                   = row.get("notes") or None,
                last_updated            = datetime.utcnow(),
            ))
            count += 1
            if count % 50 == 0:
                db.session.flush()
    db.session.commit()
    print(f"Seeded {count} companies from CSV")


app = create_app()

# ── Background scrape state ──────────────────────────────────────────────────

_scrape_lock    = threading.Lock()
_scrape_running = False
_scrape_log: list[str] = []
_scrape_counts: dict   = {}


# ── Template filters & helpers ───────────────────────────────────────────────

@app.template_filter("format_dt")
def format_dt(value):
    if not value:
        return "—"
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value)
        except ValueError:
            return value
    return value.strftime("%-m/%-d/%y")


@app.template_filter("status_label")
def status_label(value):
    return dict(PIPELINE_STATUSES).get(value, value or "—")


@app.template_filter("type_label")
def type_label(value):
    return dict(COMPANY_TYPES).get(value, value or "—")


@app.template_filter("setting_label")
def setting_label(value):
    return dict(DELIVERY_SETTINGS).get(value, value or "—")


@app.context_processor
def inject_globals():
    return {
        "company_types":     COMPANY_TYPES,
        "delivery_settings": DELIVERY_SETTINGS,
        "pipeline_statuses": PIPELINE_STATUSES,
        "revenue_estimates": REVENUE_ESTIMATES,
        "funding_statuses":  FUNDING_STATUSES,
        "channels":          CHANNELS,
        "outcomes":          OUTCOMES,
        "priority_states":   PRIORITY_STATES,
        "today":             date.today(),
    }


# ── Dashboard ────────────────────────────────────────────────────────────────

@app.route("/")
def dashboard():
    total     = Company.query.count()
    high_rank = Company.query.filter(Company.rank_score >= 70).count()
    flagged   = Company.query.filter_by(status="flagged_for_review").count()
    contacted = Company.query.filter(Company.status.in_(
        ["contacted", "responded", "meeting_scheduled"])).count()

    state_counts = dict(
        db.session.query(Company.state, func.count(Company.id))
        .group_by(Company.state).all()
    )

    state_grid = []
    for name, abbr, rank in PRIORITY_STATES:
        cnt = state_counts.get(abbr, 0)
        state_grid.append({"name": name, "abbr": abbr, "count": cnt, "rank": rank})

    top5 = sorted(state_grid, key=lambda x: x["count"], reverse=True)[:5]

    top10 = (
        Company.query
        .filter(Company.acquirable == True)
        .order_by(Company.rank_score.desc())
        .limit(10)
        .all()
    )

    recent = (
        OutreachLog.query
        .order_by(OutreachLog.date.desc())
        .limit(8)
        .all()
    )

    return render_template(
        "dashboard.html",
        total=total,
        high_rank=high_rank,
        flagged=flagged,
        contacted=contacted,
        state_grid=state_grid,
        top5=top5,
        top10=top10,
        recent=recent,
        scrape_running=_scrape_running,
    )


# ── Companies list ────────────────────────────────────────────────────────────

@app.route("/companies")
def companies():
    page        = max(1, int(request.args.get("page", 1)))
    per_page    = 50
    state_f     = request.args.get("state", "")
    type_f      = request.args.get("type", "")
    setting_f   = request.args.get("setting", "")
    score_f     = request.args.get("score", "")
    acquirable_f= request.args.get("acquirable", "")
    status_f    = request.args.get("status", "")
    search      = request.args.get("search", "").strip()
    sort        = request.args.get("sort", "rank_score")
    direction   = request.args.get("dir", "desc")

    q = Company.query

    if state_f:
        q = q.filter(Company.state == state_f)
    if type_f:
        q = q.filter(Company.company_type == type_f)
    if setting_f:
        q = q.filter(Company.delivery_setting == setting_f)
    if score_f == "high":
        q = q.filter(Company.rank_score >= 70)
    elif score_f == "mid":
        q = q.filter(Company.rank_score >= 50, Company.rank_score < 70)
    elif score_f == "low":
        q = q.filter(Company.rank_score < 50)
    if acquirable_f == "yes":
        q = q.filter(Company.acquirable == True)
    elif acquirable_f == "no":
        q = q.filter(Company.acquirable == False)
    if status_f and status_f != "all":
        q = q.filter(Company.status == status_f)
    if search:
        like = f"%{search}%"
        q = q.filter(db.or_(
            Company.company_name.ilike(like),
            Company.city.ilike(like),
            Company.description.ilike(like),
            Company.email.ilike(like),
        ))

    sortable = {"rank_score", "company_name", "state", "status", "year_founded"}
    if sort not in sortable:
        sort = "rank_score"
    sort_col = getattr(Company, sort)
    if sort == "company_name":
        sort_col = func.lower(Company.company_name)
    q = q.order_by(sort_col.desc() if direction == "desc" else sort_col.asc())

    pagination = q.paginate(page=page, per_page=per_page, error_out=False)

    db_states = sorted(set(
        r[0] for r in db.session.query(Company.state)
        .filter(Company.state != None, Company.state != "").distinct().all()
    ))

    return render_template(
        "companies.html",
        companies=pagination.items,
        pagination=pagination,
        db_states=db_states,
        state_f=state_f,
        type_f=type_f,
        setting_f=setting_f,
        score_f=score_f,
        acquirable_f=acquirable_f,
        status_f=status_f,
        search=search,
        sort=sort,
        direction=direction,
    )


# ── Company detail ────────────────────────────────────────────────────────────

@app.route("/companies/<int:company_id>")
def company_detail(company_id):
    company = Company.query.get_or_404(company_id)
    logs = (
        OutreachLog.query
        .filter_by(company_id=company_id)
        .order_by(OutreachLog.date.desc())
        .all()
    )
    return render_template("company_detail.html", company=company, logs=logs)


# ── Add company ───────────────────────────────────────────────────────────────

@app.route("/companies/add", methods=["GET", "POST"])
def company_add():
    if request.method == "POST":
        f = request.form
        score_raw = f.get("rank_score", "50").strip()
        try:
            score = max(1, min(100, int(score_raw)))
        except ValueError:
            score = 50

        company = Company(
            company_name            = f.get("company_name", "").strip(),
            website_url             = f.get("website_url", "").strip() or None,
            phone                   = f.get("phone", "").strip() or None,
            email                   = f.get("email", "").strip() or None,
            address                 = f.get("address", "").strip() or None,
            city                    = f.get("city", "").strip() or None,
            state                   = f.get("state", "").strip() or None,
            zip                     = f.get("zip", "").strip() or None,
            description             = f.get("description", "").strip() or None,
            company_type            = f.get("company_type", "behavioral_health"),
            delivery_setting        = f.get("delivery_setting", "multi_setting"),
            employee_count_estimate = f.get("employee_count_estimate", "").strip() or None,
            year_founded            = f.get("year_founded", "").strip() or None,
            revenue_estimate        = f.get("revenue_estimate", "unknown"),
            funding_status          = f.get("funding_status", "unknown"),
            acquirable              = f.get("acquirable") != "no",
            rank_score              = score,
            status                  = f.get("status", "uncontacted"),
            verified                = f.get("verified") == "on",
            notes                   = f.get("notes", "").strip() or None,
            last_updated            = datetime.utcnow(),
        )
        db.session.add(company)
        db.session.flush()

        if f.get("owner_name"):
            contact = Contact(
                company_id   = company.id,
                name         = f.get("owner_name", "").strip(),
                title        = f.get("owner_title", "").strip(),
                phone        = f.get("owner_phone", "").strip() or None,
                email        = f.get("owner_email", "").strip() or None,
                linkedin_url = f.get("owner_linkedin", "").strip() or None,
                is_primary   = True,
            )
            db.session.add(contact)

        db.session.commit()
        flash(f'"{company.company_name}" added successfully.', "success")
        return redirect(url_for("company_detail", company_id=company.id))

    return render_template("company_add.html")


# ── Flagged queue ─────────────────────────────────────────────────────────────

@app.route("/flagged")
def flagged():
    items = (
        Company.query
        .filter_by(status="flagged_for_review")
        .order_by(Company.rank_score.desc())
        .all()
    )
    return render_template("flagged.html", companies=items)


@app.route("/flagged/<int:company_id>/approve", methods=["POST"])
def flagged_approve(company_id):
    company = Company.query.get_or_404(company_id)
    company.acquirable  = True
    company.status      = "uncontacted"
    company.last_updated = datetime.utcnow()
    db.session.commit()
    flash(f'"{company.company_name}" approved and moved to target list.', "success")
    return redirect(url_for("flagged"))


@app.route("/flagged/<int:company_id>/reject", methods=["POST"])
def flagged_reject(company_id):
    company = Company.query.get_or_404(company_id)
    name = company.company_name
    db.session.delete(company)
    db.session.commit()
    flash(f'"{name}" removed from the database.', "info")
    return redirect(url_for("flagged"))


# ── Outreach tracker ──────────────────────────────────────────────────────────

@app.route("/outreach")
def outreach():
    channel_f = request.args.get("channel", "")
    outcome_f = request.args.get("outcome", "")
    from_f    = request.args.get("from", "")
    to_f      = request.args.get("to", "")

    q = OutreachLog.query.order_by(OutreachLog.date.desc())
    if channel_f:
        q = q.filter(OutreachLog.channel == channel_f)
    if outcome_f:
        q = q.filter(OutreachLog.outcome == outcome_f)
    if from_f:
        try:
            q = q.filter(OutreachLog.date >= datetime.strptime(from_f, "%Y-%m-%d"))
        except ValueError:
            pass
    if to_f:
        try:
            q = q.filter(OutreachLog.date <= datetime.strptime(to_f, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass

    logs = q.all()

    followups = (
        OutreachLog.query
        .filter(OutreachLog.follow_up_date != None)
        .filter(OutreachLog.follow_up_date <= date.today())
        .order_by(OutreachLog.follow_up_date.asc())
        .all()
    )

    all_companies = Company.query.order_by(Company.company_name).all()

    return render_template(
        "outreach.html",
        logs=logs,
        followups=followups,
        all_companies=all_companies,
        channel_f=channel_f,
        outcome_f=outcome_f,
        from_f=from_f,
        to_f=to_f,
    )


# ── Export page ───────────────────────────────────────────────────────────────

@app.route("/export")
def export_page():
    return render_template("export.html")


# ── API: company PATCH ────────────────────────────────────────────────────────

@app.route("/api/companies/<int:company_id>", methods=["PATCH"])
def api_company_patch(company_id):
    company = Company.query.get_or_404(company_id)
    data    = request.get_json(silent=True) or {}

    allowed = {
        "company_name", "website_url", "phone", "email",
        "address", "city", "state", "zip", "description",
        "company_type", "delivery_setting", "employee_count_estimate",
        "year_founded", "revenue_estimate", "funding_status",
        "acquirable", "rank_score", "status", "verified",
        "positive_signals", "exclude_flags", "notes",
    }
    for field, value in data.items():
        if field in allowed:
            if field in ("verified", "acquirable"):
                setattr(company, field, bool(value))
            elif field == "rank_score":
                try:
                    setattr(company, field, max(1, min(100, int(value))))
                except (ValueError, TypeError):
                    pass
            else:
                setattr(company, field, value)

    company.last_updated = datetime.utcnow()
    db.session.commit()
    return jsonify(company.to_dict())


# ── API: contacts ─────────────────────────────────────────────────────────────

@app.route("/api/companies/<int:company_id>/contacts", methods=["GET"])
def api_contacts_list(company_id):
    Company.query.get_or_404(company_id)
    contacts = Contact.query.filter_by(company_id=company_id).all()
    return jsonify([c.to_dict() for c in contacts])


@app.route("/api/companies/<int:company_id>/contacts", methods=["POST"])
def api_contact_add(company_id):
    Company.query.get_or_404(company_id)
    data = request.get_json(silent=True) or {}

    if data.get("is_primary"):
        Contact.query.filter_by(company_id=company_id, is_primary=True).update({"is_primary": False})

    contact = Contact(
        company_id   = company_id,
        name         = data.get("name", ""),
        title        = data.get("title", ""),
        phone        = data.get("phone", "") or None,
        email        = data.get("email", "") or None,
        linkedin_url = data.get("linkedin_url", "") or None,
        is_primary   = bool(data.get("is_primary", False)),
        notes        = data.get("notes", "") or None,
    )
    db.session.add(contact)
    db.session.commit()
    return jsonify(contact.to_dict()), 201


@app.route("/api/contacts/<int:contact_id>", methods=["PATCH"])
def api_contact_patch(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    data    = request.get_json(silent=True) or {}

    if data.get("is_primary"):
        Contact.query.filter_by(company_id=contact.company_id, is_primary=True).update({"is_primary": False})

    for field in ("name", "title", "phone", "email", "linkedin_url", "notes", "is_primary"):
        if field in data:
            setattr(contact, field, data[field])

    db.session.commit()
    return jsonify(contact.to_dict())


@app.route("/api/contacts/<int:contact_id>", methods=["DELETE"])
def api_contact_delete(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    db.session.delete(contact)
    db.session.commit()
    return jsonify({"ok": True})


# ── API: outreach log ─────────────────────────────────────────────────────────

@app.route("/api/outreach", methods=["POST"])
def api_outreach_add():
    data = request.get_json(silent=True) or {}

    company_id = data.get("company_id")
    if not company_id:
        return jsonify({"error": "company_id required"}), 400

    follow_up = None
    if data.get("follow_up_date"):
        try:
            follow_up = datetime.strptime(data["follow_up_date"], "%Y-%m-%d").date()
        except ValueError:
            pass

    log = OutreachLog(
        company_id     = company_id,
        contact_id     = data.get("contact_id") or None,
        date           = datetime.utcnow(),
        channel        = data.get("channel", "email"),
        outcome        = data.get("outcome", "no_response"),
        notes          = data.get("notes", "") or None,
        follow_up_date = follow_up,
    )
    db.session.add(log)

    company = Company.query.get(company_id)
    if company:
        if log.outcome == "meeting_set":
            company.status = "meeting_scheduled"
        elif log.outcome in ("positive", "no_response") and company.status == "uncontacted":
            company.status = "contacted"
        elif log.outcome == "negative":
            company.status = "passed"
        company.last_updated = datetime.utcnow()

    db.session.commit()
    return jsonify(log.to_dict()), 201


@app.route("/api/outreach/<int:log_id>", methods=["DELETE"])
def api_outreach_delete(log_id):
    log = OutreachLog.query.get_or_404(log_id)
    db.session.delete(log)
    db.session.commit()
    return jsonify({"ok": True})


# ── API: find owner via LinkedIn ──────────────────────────────────────────────

@app.route("/api/companies/<int:company_id>/find-owner", methods=["POST"])
def api_find_owner(company_id):
    company = Company.query.get_or_404(company_id)
    try:
        from scraper import find_owner_linkedin
        urls = find_owner_linkedin(company.company_name, company.state or "")
        return jsonify({"urls": urls})
    except Exception as e:
        return jsonify({"error": str(e), "urls": []}), 200


# ── Export: CSV ───────────────────────────────────────────────────────────────

@app.route("/api/export/csv")
def export_csv():
    state_f  = request.args.get("state", "")
    status_f = request.args.get("status", "")
    score_f  = request.args.get("score", "")

    q = Company.query
    if state_f:
        q = q.filter(Company.state == state_f)
    if status_f:
        q = q.filter(Company.status == status_f)
    if score_f == "high":
        q = q.filter(Company.rank_score >= 70)

    companies_list = q.order_by(Company.rank_score.desc()).all()

    fieldnames = [
        "company_name", "company_type", "delivery_setting", "address", "city",
        "state", "zip", "phone", "email", "website_url", "description",
        "revenue_estimate", "rank_score", "positive_signals", "status",
        "verified", "notes",
    ]

    type_map    = dict(COMPANY_TYPES)
    setting_map = dict(DELIVERY_SETTINGS)
    status_map  = dict(PIPELINE_STATUSES)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for co in companies_list:
        writer.writerow({
            "company_name":    co.company_name or "",
            "company_type":    type_map.get(co.company_type, co.company_type or ""),
            "delivery_setting":setting_map.get(co.delivery_setting, co.delivery_setting or ""),
            "address":         co.address or "",
            "city":            co.city or "",
            "state":           (co.state or "").upper(),
            "zip":             co.zip or "",
            "phone":           co.phone or "",
            "email":           (co.email or "").lower(),
            "website_url":     co.website_url or "",
            "description":     (co.description or "").replace("\n", " "),
            "revenue_estimate":co.revenue_estimate or "",
            "rank_score":      co.rank_score or "",
            "positive_signals":co.positive_signals or "",
            "status":          status_map.get(co.status, co.status or ""),
            "verified":        "Yes" if co.verified else "No",
            "notes":           (co.notes or "").replace("\n", " "),
        })

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=bh_targets.csv"},
    )


# ── Export: Excel ─────────────────────────────────────────────────────────────

@app.route("/api/export/excel")
def export_excel():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    state_f  = request.args.get("state", "")
    status_f = request.args.get("status", "")

    q = Company.query
    if state_f:
        q = q.filter(Company.state == state_f)
    if status_f:
        q = q.filter(Company.status == status_f)
    companies_list = q.order_by(Company.rank_score.desc()).all()

    wb = openpyxl.Workbook()
    navy_fill   = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    gray_fill   = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)

    def _style(ws, headers):
        ws.freeze_panes = "A2"
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = navy_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 20

    def _autowidth(ws):
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 55)

    def _shade(ws, n):
        for i in range(2, n + 2):
            if i % 2 == 0:
                for cell in ws[i]:
                    cell.fill = gray_fill

    type_map    = dict(COMPANY_TYPES)
    setting_map = dict(DELIVERY_SETTINGS)
    status_map  = dict(PIPELINE_STATUSES)
    rev_map     = dict(REVENUE_ESTIMATES)
    fund_map    = dict(FUNDING_STATUSES)

    # Sheet 1: Companies
    ws1 = wb.active
    ws1.title = "Targets"
    h1 = ["Company", "Type", "Setting", "City", "State", "Phone", "Email", "Website",
          "Description", "Revenue Est", "Rank Score", "Funding", "Positive Signals",
          "Status", "Verified", "Notes", "Last Updated"]
    _style(ws1, h1)
    for co in companies_list:
        pc = co.primary_contact
        ws1.append([
            co.company_name,
            type_map.get(co.company_type, co.company_type or ""),
            setting_map.get(co.delivery_setting, co.delivery_setting or ""),
            co.city or "",
            co.state or "",
            co.phone or "",
            co.email or "",
            co.website_url or "",
            co.description or "",
            rev_map.get(co.revenue_estimate, co.revenue_estimate or ""),
            co.rank_score or 0,
            fund_map.get(co.funding_status, co.funding_status or ""),
            co.positive_signals or "",
            status_map.get(co.status, co.status or ""),
            "Yes" if co.verified else "No",
            co.notes or "",
            co.last_updated.strftime("%Y-%m-%d") if co.last_updated else "",
        ])
    _shade(ws1, len(companies_list))
    _autowidth(ws1)

    # Sheet 2: Contacts
    ws2 = wb.create_sheet("Contacts")
    _style(ws2, ["Company", "Name", "Title", "Phone", "Email", "LinkedIn", "Primary"])
    all_contacts = Contact.query.all()
    for ct in all_contacts:
        co = Company.query.get(ct.company_id)
        ws2.append([
            co.company_name if co else "",
            ct.name or "", ct.title or "", ct.phone or "",
            ct.email or "", ct.linkedin_url or "",
            "Yes" if ct.is_primary else "No",
        ])
    _shade(ws2, len(all_contacts))
    _autowidth(ws2)

    # Sheet 3: Outreach Log
    ws3 = wb.create_sheet("Outreach Log")
    _style(ws3, ["Company", "Contact", "Date", "Channel", "Outcome", "Follow-Up", "Notes"])
    ch_map  = dict(CHANNELS)
    out_map = dict(OUTCOMES)
    for lg in OutreachLog.query.order_by(OutreachLog.date.desc()).all():
        co = Company.query.get(lg.company_id)
        ct = Contact.query.get(lg.contact_id) if lg.contact_id else None
        ws3.append([
            co.company_name if co else "",
            ct.name if ct else "",
            lg.date.strftime("%Y-%m-%d") if lg.date else "",
            ch_map.get(lg.channel, lg.channel or ""),
            out_map.get(lg.outcome, lg.outcome or ""),
            lg.follow_up_date.strftime("%Y-%m-%d") if lg.follow_up_date else "",
            lg.notes or "",
        ])
    _shade(ws3, OutreachLog.query.count())
    _autowidth(ws3)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bh_targets.xlsx"},
    )


# ── Scrape: run discovery ─────────────────────────────────────────────────────

@app.route("/scrape/run", methods=["POST"])
def scrape_run():
    global _scrape_running, _scrape_log, _scrape_counts

    data   = request.get_json(silent=True) or {}
    states = data.get("states", [])

    if not states:
        return jsonify({"error": "No states provided"}), 400

    with _scrape_lock:
        if _scrape_running:
            return jsonify({"error": "A scrape is already running"}), 409
        _scrape_running = True
        _scrape_log     = []
        _scrape_counts  = {}

    def run():
        global _scrape_running, _scrape_log, _scrape_counts
        try:
            from scraper import discover_companies_by_state
            for abbr in states:
                _scrape_log.append(f"Scanning {abbr}…")
                try:
                    new_cos = discover_companies_by_state(abbr, None, app)
                    _scrape_counts[abbr] = len(new_cos)
                    _scrape_log.append(f"  → {len(new_cos)} new companies in {abbr}")
                except Exception as exc:
                    _scrape_log.append(f"  ✗ Error in {abbr}: {exc}")
        except Exception as e:
            _scrape_log.append(f"Fatal error: {e}")
        finally:
            _scrape_running = False
            _scrape_log.append("Done.")

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"status": "started", "states": states})


@app.route("/scrape/status")
def scrape_status():
    return jsonify({
        "running": _scrape_running,
        "log":     _scrape_log[-200:],
        "counts":  _scrape_counts,
    })


# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5100, debug=True, use_reloader=False)
